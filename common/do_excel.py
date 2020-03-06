#!usr/bin/env python
#-*- coding:utf-8 _*-
"""
@author:末夏
@file: do_excel.py
@time: 2019/08/19
"""
from openpyxl import load_workbook
import os
from common import constant
from common.config import Config
import json

class Case:
    def __init__(self):
        self.sheet_name=None
        self.case_id=None
        self.title=None
        self.url=None
        self.data=None
        self.method=None
        self.expect=None
        self.actual=None
        self.result=None
class DoExcel:
    def __init__(self,filename):
        try:
            self.filename=filename                         #测试用例excel文件名字
            self.file_name=load_workbook(self.filename)    #加载excel工作簿
            cast_list=Config().get_value('case','case_list')
            self.run_sheet=eval(cast_list)        #需要执行的excel表格,在配置文件中设置
        except Exception as e:
            print('打开excel文件异常')

    def read_all_excel(self):                              #读取excel数据
        sheets=self.get_sheets()                        #获取该文件的所有工作表
        for sheet_name in sheets:
            try:#循环读取每一个表格
                if sheet_name in self.run_sheet:
                    sheet1=self.file_name[sheet_name]
                    maxrow=sheet1.max_row
                    cases=[]
                    for r in range(2,maxrow+1):                #读取每一行的单元格数据 ，实例化case对象，
                        case = Case()
                        case.sheet_name=sheet_name
                        case.case_id=sheet1.cell(r,1).value
                        case.title = sheet1.cell(r, 2).value
                        case.url = sheet1.cell(r, 3).value
                        case.data = sheet1.cell(r, 4).value
                        case.method = sheet1.cell(r, 5).value
                        case.expect = sheet1.cell(r, 6).value
                        case.actual = sheet1.cell(r, 7).value
                        case.result = sheet1.cell(r, 8).value
                        cases.append(case)                             #把每一个实例加到 cases列表中
                    return cases
            except Exception as e:
                print('读取excel文件数据有误')
                raise e
    def read_excel(self,sheetname):
        sheet1 = self.file_name[sheetname]
        maxrow = sheet1.max_row
        cases = []
        for r in range(2, maxrow + 1):  # 读取每一行的单元格数据 ，实例化case对象，
            case = Case()
            case.sheet_name = sheetname
            case.case_id = sheet1.cell(r, 1).value
            case.title = sheet1.cell(r, 2).value
            case.url = sheet1.cell(r, 3).value
            case.data = sheet1.cell(r, 4).value
            case.method = sheet1.cell(r, 5).value
            case.expect = sheet1.cell(r, 6).value
            case.actual = sheet1.cell(r, 7).value
            case.result = sheet1.cell(r, 8).value
            cases.append(case)  # 把每一个实例加到 cases列表中
        return cases


    def get_sheets(self):                                         #获取该文件所有表格名字
        return self.file_name.sheetnames
    def write_all_excel(self,row,column,result):
        sheets = self.get_sheets()
        for sheet_name in sheets:
            cases=self.read_all_excel()
            for case in cases:
                if case.sheet_name==sheet_name and int(case.case_id)+1==int(row):
                    sheet1=self.file_name[sheet_name]
                    sheet1.cell(int(row), int(column)).value = result
                    self.file_name.save(self.filename)

    def write_excel(self,sheetname,row,column,result):
        cases = self.read_excel(sheetname)
        for case in cases:
            if case.sheet_name == sheetname and int(case.case_id) + 1 == int(row):
                sheet1 = self.file_name[sheetname]
                sheet1.cell(int(row), int(column)).value = result
                self.file_name.save(self.filename)





    # def write_excel(self,sheetname,row,column,result):             #写入数据到excell表中
    #
    #     sheet1=self.file_name[sheetname]
    #     sheet1.cell(int(row),int(column)).value=result
    #     self.file_name.save(self.filename)

# if __name__=="__main__":
#     cases=DoExcel(os.path.join(constant.datas_path,'data.xlsx')).read_excel()
#     for case in cases:
#         print(case.title)
#         print(case.data)





